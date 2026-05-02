"""
artifacts.py
============
CRUD + download + session-load endpoints for persisted pipeline artifacts.

All endpoints are user-scoped (filter by user_id from JWT).
Presigned URLs are generated fresh on every response — never stored.
"""

import io
import logging
import time
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..db import get_db
from ..dependencies import CurrentUser, get_current_user
from ..models_db import ArtifactDB, DatasetMetaDB, DatasetDataDB, DatasetChunkDB
from ..services.object_storage import StorageService
from ..services.duckdb_session import register_view, execute_in_session
from ..services.rate_limiter import limiter

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/artifacts", tags=["artifacts"])


# ── helpers ──────────────────────────────────────────────────────────────────

def _get_artifact_or_404(artifact_id: str, user_id: str, db: Session) -> ArtifactDB:
    row = (
        db.query(ArtifactDB)
        .filter(ArtifactDB.id == artifact_id, ArtifactDB.user_id == user_id)
        .first()
    )
    if not row:
        raise HTTPException(status_code=404, detail="Artifact not found")
    return row


def _serialize(artifact: ArtifactDB, download_url: Optional[str] = None, dataset_id: Optional[str] = None) -> dict:
    return {
        "id": artifact.id,
        "name": artifact.name,
        "description": artifact.description,
        "type": artifact.type,
        "format": artifact.format,
        "row_count": artifact.row_count,
        "column_schema": artifact.column_schema or [],
        "pipeline_run_id": artifact.pipeline_run_id,
        "step_id": artifact.step_id,
        "session_id": artifact.session_id,
        "created_at": artifact.created_at.isoformat() if artifact.created_at else None,
        "download_url": download_url,
        "dataset_id": dataset_id,
    }


# ── 1. List artifacts ─────────────────────────────────────────────────────────

@router.get("")
def list_artifacts(
    session_id: Optional[str] = Query(None),
    pipeline_run_id: Optional[str] = Query(None),
    project_id: Optional[str] = Query(None),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(ArtifactDB).filter(ArtifactDB.user_id == current_user.id)
    if session_id:
        q = q.filter(ArtifactDB.session_id == session_id)
    if pipeline_run_id:
        q = q.filter(ArtifactDB.pipeline_run_id == pipeline_run_id)

    artifacts = q.order_by(ArtifactDB.created_at.desc()).all()

    # Bulk-resolve dataset_id via storage_path == s3_key
    s3_keys = [art.s3_key for art in artifacts if art.s3_key]
    datasets_by_key: dict[str, str] = {}
    dataset_project_by_key: dict[str, Optional[str]] = {}
    if s3_keys:
        ds_rows = (
            db.query(
                DatasetMetaDB.id,
                DatasetMetaDB.storage_path,
                DatasetMetaDB.project_id,
                DatasetMetaDB.deleted_at,
            )
            .filter(DatasetMetaDB.storage_path.in_(s3_keys))
            .all()
        )
        for row in ds_rows:
            # Skip soft-deleted datasets so the artifact looks orphaned
            # (and gets filtered out below when project_id is requested).
            if row.deleted_at is not None:
                continue
            datasets_by_key[row.storage_path] = str(row.id)
            dataset_project_by_key[row.storage_path] = (
                str(row.project_id) if row.project_id else None
            )

    result = []
    for art in artifacts:
        # ── Project scoping ──────────────────────────────────────────────
        # When a project_id filter is supplied, only return artifacts whose
        # underlying dataset still exists AND belongs to that project.
        # This stops artifacts from a deleted project leaking into a brand-new
        # project (the artifact rows themselves are user-scoped, not project-
        # scoped, in the schema — so we filter via the linked dataset).
        if project_id is not None:
            ds_project = dataset_project_by_key.get(art.s3_key) if art.s3_key else None
            if ds_project != project_id:
                continue

        download_url: Optional[str] = None
        if art.s3_key and not str(art.s3_key).startswith("local://"):
            try:
                download_url = StorageService.get_signed_url(art.s3_key, expires_in=3600)
            except Exception as exc:
                logger.warning("Could not generate signed URL for artifact %s: %s", art.id, exc)
        dataset_id = datasets_by_key.get(art.s3_key) if art.s3_key else None
        result.append(_serialize(art, download_url, dataset_id=dataset_id))

    return result


# ── 2. Rename artifact ────────────────────────────────────────────────────────

@router.patch("/{artifact_id}")
def rename_artifact(
    artifact_id: str,
    body: dict,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    artifact = _get_artifact_or_404(artifact_id, current_user.id, db)
    if "name" in body:
        artifact.name = str(body["name"]).strip() or artifact.name
    if "description" in body:
        artifact.description = body["description"]
    db.commit()
    db.refresh(artifact)
    return _serialize(artifact)


# ── 3. Download artifact (csv / xlsx / parquet) ───────────────────────────────

@router.get("/{artifact_id}/download")
@limiter.limit("30/hour")
def download_artifact(
    request: Request,
    artifact_id: str,
    fmt: str = Query("parquet", description="Output format: csv | xlsx | parquet"),
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    artifact = _get_artifact_or_404(artifact_id, current_user.id, db)
    s3_key = artifact.s3_key
    safe_name = (artifact.name or "artifact").replace(" ", "_")

    # Reject legacy bad-format sentinels ("local/..." without ://)
    if s3_key and str(s3_key).startswith("local/") and "://" not in str(s3_key):
        raise HTTPException(
            status_code=503,
            detail="This artifact was generated without cloud storage configured and cannot be downloaded.",
        )

    if fmt == "parquet":
        if s3_key and str(s3_key).startswith("local://"):
            # Local artifacts must be served directly, not via redirect
            try:
                raw_bytes = _fetch_bytes(s3_key)
            except Exception as exc:
                raise HTTPException(status_code=500, detail=f"Failed to read local artifact: {exc}")
            return Response(
                content=raw_bytes,
                media_type="application/octet-stream",
                headers={"Content-Disposition": f'attachment; filename="{safe_name}.parquet"'},
            )
        # Return a redirect to a fresh presigned URL
        try:
            url = StorageService.get_signed_url(s3_key, expires_in=3600)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Could not generate download URL: {exc}")
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=url)

    # Download parquet bytes from storage and convert
    try:
        raw_bytes = _fetch_bytes(s3_key)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to fetch artifact from storage: {exc}")

    import pyarrow.parquet as _pq
    _tbl = _pq.read_table(io.BytesIO(raw_bytes))
    df = _tbl.to_pandas()

    if fmt == "csv":
        buf = io.StringIO()
        df.to_csv(buf, index=False)
        content = buf.getvalue().encode("utf-8")
        return Response(
            content=content,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'},
        )

    if fmt in ("xlsx", "excel"):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = safe_name[:31]
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        cols = list(df.columns)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = header_fill
            cell.font = header_font
        for ri, row_data in enumerate(df.itertuples(index=False), 2):
            for ci, val in enumerate(row_data, 1):
                ws.cell(row=ri, column=ci, value=val)
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
        )

    raise HTTPException(status_code=400, detail=f"Unsupported format '{fmt}'. Use csv, xlsx, or parquet.")


def _fetch_bytes(s3_key: str) -> bytes:
    """Download raw bytes from storage given an s3_key (storage path)."""
    if str(s3_key).startswith("local://"):
        local_path = StorageService._local_dir() / str(s3_key).replace("local://", "", 1)
        return local_path.read_bytes()
    url = StorageService.get_signed_url(s3_key, expires_in=3600)
    import urllib.request
    with urllib.request.urlopen(url, timeout=60) as resp:  # type: ignore
        return resp.read()


def _upload_with_retry(
    user_id: str,
    session_id: str,
    checkpoint_id: str,
    parquet_bytes: bytes,
) -> str:
    """Upload Parquet bytes to object storage with exponential backoff.

    Retries up to 3 times (delays: 1 s, 2 s, 4 s) before raising.
    The session VIEW remains intact throughout so users can retry the Save.
    """
    file_name = f"{checkpoint_id}.parquet"
    last_exc: Exception | None = None
    for attempt, delay in enumerate([0, 1, 2, 4]):
        if delay:
            time.sleep(delay)
        try:
            return StorageService.upload(
                user_id=user_id,
                dataset_id=f"checkpoints/{session_id}",
                buffer=parquet_bytes,
                file_name=file_name,
            )
        except Exception as exc:
            last_exc = exc
            logger.warning("S3_UPLOAD_RETRY: attempt %d failed for checkpoint %s: %s", attempt + 1, checkpoint_id, exc)
    raise last_exc  # type: ignore[misc]


# ── 4. Save pipeline step as a checkpoint artifact ───────────────────────────

@router.post("/save-checkpoint")
@limiter.limit("20/hour")
def save_checkpoint(
    request: Request,
    body: dict,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Materialise a DuckDB session table as a Parquet artifact + DatasetMetaDB row.

    Body:
        session_id         (required) — DuckDB session containing the table
        table_name         (required) — DuckDB table/view name to snapshot
        artifact_name      (optional) — human-visible label (defaults to table_name)
        description        (optional)
        source_dataset_id  (optional) — raw dataset this artifact was derived from;
                                        stored as parent_id to link lineage
    """
    session_id = str(body.get("session_id") or "").strip()
    table_name = str(body.get("table_name") or "").strip()
    if not session_id:
        raise HTTPException(status_code=422, detail="session_id is required")
    if not table_name:
        raise HTTPException(status_code=422, detail="table_name is required")

    artifact_name = str(body.get("artifact_name") or table_name).strip()
    description = body.get("description")
    source_dataset_id = str(body.get("source_dataset_id") or "").strip() or None

    # 1. Fetch rows from the live DuckDB session
    try:
        rows = execute_in_session(session_id, f"SELECT * FROM {table_name}") or []
    except Exception as exc:
        raise HTTPException(
            status_code=404,
            detail=f"Table '{table_name}' not found in session '{session_id}': {exc}",
        )

    # 2. Serialise to Parquet
    try:
        import io as _io
        import pyarrow as _pa
        import pyarrow.parquet as _pq

        if rows:
            parquet_table = _pa.Table.from_pylist(rows)
        else:
            parquet_table = _pa.table({})
        buf = _io.BytesIO()
        _pq.write_table(parquet_table, buf)
        parquet_bytes = buf.getvalue()
        col_schema: list = list(rows[0].keys()) if rows else []
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Parquet serialisation failed: {exc}")

    # 3. Upload to object storage (with exponential-backoff retry)
    import uuid as _uuid
    checkpoint_id = str(_uuid.uuid4())
    try:
        s3_key = _upload_with_retry(
            user_id=current_user.id,
            session_id=session_id,
            checkpoint_id=checkpoint_id,
            parquet_bytes=parquet_bytes,
        )
    except Exception as exc:
        raise HTTPException(
            status_code=503,
            detail=(
                f"Save failed after 3 attempts — your session is still active, "
                f"try again in a moment. ({exc})"
            ),
        )

    # 4. Persist DatasetMetaDB so dataset appears in the datasets list
    from ..services.persistence_policy import materialize_artifact, materialize_dataset
    ds_id = str(_uuid.uuid4())
    # Inherit the source dataset's project so the saved artifact stays scoped
    # to the same project the user is working in. Without this, list_artifacts
    # filters the saved artifact out (project_id NULL != active project_id).
    # SECURITY: scope the lookup to the requesting user so a malicious caller
    # cannot pass another user's dataset_id and have their artifact silently
    # adopted into the victim's project.
    source_project_id = None
    if source_dataset_id:
        src_meta = (
            db.query(DatasetMetaDB)
            .filter(
                DatasetMetaDB.id == source_dataset_id,
                DatasetMetaDB.user_id == current_user.id,
            )
            .first()
        )
        if src_meta is None:
            raise HTTPException(
                status_code=404,
                detail="source_dataset_id not found",
            )
        source_project_id = getattr(src_meta, "project_id", None)
    try:
        ds_row = materialize_dataset(
            db,
            triggered_by="user_save",
            id=ds_id,
            user_id=current_user.id,
            workspace_id="default",
            project_id=source_project_id,
            name=artifact_name,
            source_type="checkpoint",
            storage_path=s3_key,
            file_format="parquet",
            columns=col_schema,
            row_count=len(rows),
            status="ready",
            parent_id=source_dataset_id,
        )
        db.flush()
    except Exception as exc:
        logger.warning("DatasetMetaDB persist failed for checkpoint %s: %s", checkpoint_id, exc)
        ds_id = None

    # 5. Persist ArtifactDB row
    artifact_row = materialize_artifact(
        db,
        triggered_by="user_save",
        id=checkpoint_id,
        user_id=current_user.id,
        session_id=session_id,
        pipeline_run_id=None,
        step_id=None,
        name=artifact_name,
        description=description,
        s3_key=s3_key,
        row_count=len(rows),
        column_schema=col_schema,
        type="checkpoint",
        format="parquet",
    )
    db.commit()

    return {
        **_serialize(artifact_row, dataset_id=ds_id),
        "dataset_id": ds_id,
        "message": f"Checkpoint '{artifact_name}' saved ({len(rows):,} rows).",
    }


# ── 5. Load artifact into a DuckDB session ────────────────────────────────────

@router.post("/{artifact_id}/load")
def load_artifact_into_session(
    artifact_id: str,
    body: dict,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Register the artifact Parquet file as a named view in an existing DuckDB session.
    Body: { "session_id": "...", "table_name": "optional_alias" }
    """
    artifact = _get_artifact_or_404(artifact_id, current_user.id, db)
    session_id = str(body.get("session_id") or "").strip()
    if not session_id:
        raise HTTPException(status_code=422, detail="session_id is required")

    table_name = str(body.get("table_name") or artifact.name or f"artifact_{artifact_id[:8]}").strip()
    table_name = table_name.replace(" ", "_")

    try:
        query_path = StorageService.get_query_path(artifact.s3_key)
        # Register as a lazy VIEW — no rows are materialised into RAM on click.
        # DuckDB only reads from the Parquet file when a query actually runs
        # against the view, preventing an OOM-kill on large artifacts.
        register_view(session_id, table_name, query_path)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to load artifact into session: {exc}")

    return {
        "session_id": session_id,
        "table_name": table_name,
        "row_count": artifact.row_count,
        "message": f"Artifact loaded as '{table_name}' in session {session_id}",
    }


# ── 6. Delete artifact ────────────────────────────────────────────────────────

@router.delete("/{artifact_id}")
def delete_artifact(
    artifact_id: str,
    current_user: CurrentUser = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    artifact = _get_artifact_or_404(artifact_id, current_user.id, db)
    s3_key = artifact.s3_key

    # Delete the linked DatasetMetaDB record created by save-checkpoint (same s3_key = storage_path)
    linked_ds = db.query(DatasetMetaDB).filter(
        DatasetMetaDB.storage_path == s3_key
    ).first()
    if linked_ds:
        db.query(DatasetDataDB).filter(DatasetDataDB.id == linked_ds.id).delete(synchronize_session=False)
        db.query(DatasetChunkDB).filter(DatasetChunkDB.dataset_id == linked_ds.id).delete(synchronize_session=False)
        db.delete(linked_ds)

    db.delete(artifact)

    # DB rows are queued for delete; attempt the storage delete and queue on
    # failure so the commit either persists row-delete + retry-row together,
    # or rolls both back -- never orphaned objects.
    from ..services.storage_cleanup import safe_storage_delete
    safe_storage_delete(s3_key, source="artifact", db=db)

    # Persistent audit row in pipeline_events (best-effort).
    try:
        from ..services.event_log import emit_event as _emit_log_event
        _emit_log_event(
            db,
            event_type="artifact_deleted",
            user_id=current_user.id,
            session_id=getattr(artifact, "session_id", None),
            run_id=getattr(artifact, "pipeline_run_id", None),
            step_id=getattr(artifact, "step_id", None),
            payload={
                "artifact_id": artifact_id,
                "name": getattr(artifact, "name", None),
                "s3_key": s3_key,
                "linked_dataset_id": getattr(linked_ds, "id", None) if linked_ds else None,
            },
        )
    except Exception:
        pass

    db.commit()
    return Response(status_code=204)
