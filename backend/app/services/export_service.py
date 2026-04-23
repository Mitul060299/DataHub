"""
ExportService
Serialises a named DuckDB session table to CSV / Excel / Parquet,
uploads the result to object storage, and returns a signed download URL.
"""
from __future__ import annotations

import io
import uuid
from datetime import datetime
from typing import Literal

import pyarrow as pa
import pyarrow.parquet as pq

from .duckdb_session import execute_in_session, get_connection
from .object_storage import StorageService
from ..db import SessionLocal
from ..models_db import DatasetMetaDB

ExportFormat = Literal["csv", "excel", "parquet"]


class ExportService:
    @staticmethod
    def export(
        session_id: str,
        duckdb_name: str,
        fmt: ExportFormat,
        dataset_id: str,
        user_id: str,
        display_name: str,
        db=None,
    ) -> str:
        """
        Query *duckdb_name* from the session DuckDB connection, serialise to
        *fmt*, upload to object storage, and return a signed download URL.
        """
        rows = execute_in_session(session_id, f"SELECT * FROM {duckdb_name}")
        if not rows:
            raise ValueError(f"Table '{duckdb_name}' is empty or does not exist in the session.")

        safe_name = display_name.replace(" ", "_").replace("/", "_")
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        file_stem = f"{safe_name}_{timestamp}"

        if fmt == "csv":
            buf = io.BytesIO()
            import csv as _csv
            wrapper = io.TextIOWrapper(buf, encoding="utf-8", newline="")
            writer = _csv.DictWriter(wrapper, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
            wrapper.flush()
            wrapper.detach()
            buf.seek(0)
            file_name = f"{file_stem}.csv"
            mime_type = "text/csv"
        elif fmt == "excel":
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            safe_sheet = safe_name[:31].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "")
            ws.title = safe_sheet or "Sheet1"
            headers = list(rows[0].keys())
            date_cols = {h for h in headers if any(kw in h.lower() for kw in ("date", "_at", "_on"))}
            variance_cols = [h for h in headers if any(kw in h.lower() for kw in ("variance", "diff", "difference"))]

            header_fill = PatternFill("solid", fgColor="D9D9D9")
            header_font = Font(bold=True)
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            red_fill = PatternFill("solid", fgColor="FFCCCC")
            green_fill = PatternFill("solid", fgColor="CCFFCC")
            for row_idx, row in enumerate(rows, start=2):
                # Determine row highlight from variance cols
                row_fill = None
                if variance_cols:
                    for vc in variance_cols:
                        vc_val = row.get(vc)
                        try:
                            is_zero = float(vc_val) == 0.0 if vc_val not in (None, "") else True
                        except (TypeError, ValueError):
                            is_zero = vc_val in (None, "", "0", 0)
                        row_fill = green_fill if is_zero else red_fill
                for col_idx, h in enumerate(headers, start=1):
                    val = row.get(h)
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    if h in date_cols and val is not None:
                        cell.number_format = "DD/MM/YYYY"
                    if row_fill:
                        cell.fill = row_fill

            ws.freeze_panes = "A2"
            for col_idx, h in enumerate(headers, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = max((len(str(r.get(h) or "")) for r in rows), default=0)
                ws.column_dimensions[col_letter].width = min(max(len(h), max_len) + 2, 40)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            date_str = datetime.utcnow().strftime("%Y-%m-%d")
            file_name = f"{safe_name}_{date_str}.xlsx"
            mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif fmt == "parquet":
            keys = list(rows[0].keys())
            col_data: dict[str, list] = {k: [r.get(k) for r in rows] for k in keys}
            table = pa.table(col_data)
            buf = io.BytesIO()
            pq.write_table(table, buf)
            buf.seek(0)
            file_name = f"{file_stem}.parquet"
            mime_type = "application/octet-stream"
        else:
            raise ValueError(f"Unsupported export format: {fmt}")

        storage_path = StorageService.upload(
            user_id=user_id,
            dataset_id=dataset_id,
            buffer=buf.read(),
            file_name=file_name,
        )

        signed_url = StorageService.get_signed_url(storage_path, expires_in=3600 * 24)
        return signed_url

    @staticmethod
    def _row_count_from_session(session_id: str, duckdb_name: str) -> int:
        try:
            rows = execute_in_session(session_id, f"SELECT COUNT(*) AS n FROM {duckdb_name}")
            return int(rows[0]["n"]) if rows else 0
        except Exception:
            return 0

    @staticmethod
    def _load_df_for_dataset(
        dataset_id: str,
        db,
        *,
        storage_path_override: str | None = None,
    ) -> "pd.DataFrame":
        """Load the full dataset as a DataFrame — uses DuckDB read_parquet on stored artifact.
        Falls back to DatasetChunkDB chunks for legacy datasets with no storage_path.

        ``storage_path_override`` lets callers point at a pipeline-step snapshot
        parquet (post-transformation) instead of the original upload, so exports
        reflect changes made by the AI agent.
        """
        import pandas as pd
        from ..models_db import DatasetChunkDB
        from .duckdb_service import DuckDBService

        meta = db.query(DatasetMetaDB).filter(DatasetMetaDB.id == dataset_id).first()
        if not meta:
            raise ValueError(f"Dataset '{dataset_id}' not found")

        effective_path = storage_path_override or meta.storage_path
        if effective_path:
            query_path = StorageService.get_query_path(effective_path)
            conn = DuckDBService._ensure_db()
            df = conn.execute(f"SELECT * FROM read_parquet('{query_path}')").df()
        else:
            # Legacy fallback: reconstruct from chunk table
            chunks = (
                db.query(DatasetChunkDB)
                .filter(DatasetChunkDB.dataset_id == dataset_id)
                .order_by(DatasetChunkDB.chunk_index.asc())
                .all()
            )
            rows: list[dict] = []
            for chunk in chunks:
                rows.extend(chunk.rows or [])
            if not rows:
                raise ValueError(f"Dataset '{dataset_id}' has no data")
            df = pd.DataFrame(rows)
        return df

    @staticmethod
    def export_powerbi(
        dataset_id: str,
        display_name: str,
        db,
        *,
        storage_path_override: str | None = None,
    ) -> bytes:
        """Export full dataset as a Power BI-ready .xlsx file (openpyxl, no LIMIT)."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        df = ExportService._load_df_for_dataset(
            dataset_id, db, storage_path_override=storage_path_override
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        safe_sheet = display_name[:31].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "")
        ws.title = safe_sheet or "Data"

        headers = list(df.columns)
        date_cols = {h for h in headers if any(kw in h.lower() for kw in ("date", "_at", "_on"))}

        header_fill = PatternFill("solid", fgColor="D9D9D9")
        header_font = Font(bold=True)
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, (h, val) in enumerate(zip(headers, row), start=1):
                # Convert numpy/pandas scalars to Python natives for openpyxl.
                # openpyxl cannot serialise pd.NaT, pd.NA, pd.Timestamp, or
                # numpy scalar types — convert all of them to plain Python objects.
                try:
                    import numpy as np
                    import pandas as _pd
                    if val is _pd.NaT or val is _pd.NA:
                        val = None
                    elif isinstance(val, _pd.Timestamp):
                        val = val.to_pydatetime()
                    elif isinstance(val, (np.integer,)):
                        val = int(val)
                    elif isinstance(val, (np.floating,)):
                        val = None if np.isnan(val) else float(val)
                    elif isinstance(val, (np.bool_,)):
                        val = bool(val)
                    elif isinstance(val, float) and val != val:  # bare NaN
                        val = None
                except ImportError:
                    pass
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if h in date_cols and val is not None:
                    cell.number_format = "DD/MM/YYYY"

        ws.freeze_panes = "A2"
        for col_idx, h in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            # Sample up to 1000 rows for column width calculation to avoid O(n) on huge datasets
            sample = df[h].astype(str).head(1000)
            max_len = sample.str.len().max() if not sample.empty else 0
            ws.column_dimensions[col_letter].width = min(max(len(h), int(max_len or 0)) + 2, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    @staticmethod
    def export_tableau(
        dataset_id: str,
        display_name: str,
        db,
        *,
        storage_path_override: str | None = None,
    ) -> tuple[bytes, str]:
        """Export full dataset as a Tableau .hyper file (pantab).
        Returns (file_bytes, mime_type).  Falls back to CSV if pantab is unavailable."""
        df = ExportService._load_df_for_dataset(
            dataset_id, db, storage_path_override=storage_path_override
        )

        try:
            import pantab
            import tableauhyperapi as tab_api

            # Map DuckDB/pandas dtypes to Tableau SqlType
            def _hyper_type(dtype_str: str):
                d = dtype_str.lower()
                if "int" in d:
                    return tab_api.SqlType.big_int()
                if "float" in d or "double" in d:
                    return tab_api.SqlType.double()
                if "bool" in d:
                    return tab_api.SqlType.bool()
                if "date" in d and "time" not in d:
                    return tab_api.SqlType.date()
                if "timestamp" in d or "datetime" in d:
                    return tab_api.SqlType.timestamp()
                return tab_api.SqlType.text()

            table_def = tab_api.TableDefinition(
                table_name=tab_api.TableName("Extract", "Extract"),
                columns=[
                    tab_api.TableDefinition.Column(
                        name=col,
                        type=_hyper_type(str(df[col].dtype)),
                        nullability=tab_api.NULLABLE,
                    )
                    for col in df.columns
                ],
            )

            buf = io.BytesIO()
            pantab.frame_to_hyper(df, buf, table=tab_api.TableName("Extract", "Extract"))
            return buf.getvalue(), "application/octet-stream"

        except Exception:
            # Fallback: return CSV
            csv_buf = io.StringIO()
            df.to_csv(csv_buf, index=False)
            return csv_buf.getvalue().encode("utf-8"), "text/csv"

    @staticmethod
    def export_sheets(
        dataset_id: str,
        spreadsheet_url: str,
        sheet_name: str,
        mode: str,
        db,
        *,
        storage_path_override: str | None = None,
    ) -> dict:
        """Sync full dataset to a Google Sheet via a service-account credential.
        mode='replace' clears the sheet first; mode='append' adds below last row.
        Returns {rows_written, spreadsheet_url, sheet_name}.
        """
        import json
        import gspread
        from google.oauth2.service_account import Credentials
        from ..config import settings

        sa_json = settings.google_service_account_json
        if not sa_json:
            raise ValueError("GOOGLE_SERVICE_ACCOUNT_JSON is not configured on this server")

        # Accept raw JSON string or base64-encoded JSON
        try:
            sa_info = json.loads(sa_json)
        except json.JSONDecodeError:
            import base64
            sa_info = json.loads(base64.b64decode(sa_json).decode("utf-8"))

        scopes = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_info(sa_info, scopes=scopes)
        gc = gspread.authorize(creds)

        df = ExportService._load_df_for_dataset(
            dataset_id, db, storage_path_override=storage_path_override
        )

        spreadsheet = gc.open_by_url(spreadsheet_url)
        try:
            ws = spreadsheet.worksheet(sheet_name)
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet(title=sheet_name, rows=str(len(df) + 10), cols=str(len(df.columns) + 2))

        if mode == "replace":
            ws.clear()
            start_row = 1
        else:
            existing = ws.get_all_values()
            start_row = len(existing) + 1  # append after last row

        # Write header only for replace (or empty sheet on append)
        rows_to_write: list[list] = []
        if mode == "replace" or start_row == 1:
            rows_to_write.append(list(df.columns))

        # Convert df rows to plain Python lists (gspread needs JSON-serialisable values)
        for _, row in df.iterrows():
            rows_to_write.append([
                (None if (hasattr(v, "__class__") and v.__class__.__name__ in ("float", "float64") and str(v) == "nan") else
                 (bool(v) if hasattr(v, "__class__") and v.__class__.__name__ == "bool_" else
                  str(v) if not isinstance(v, (int, float, str, bool, type(None))) else v))
                for v in row
            ])

        # Batch write in 10K-row chunks to respect Sheets API limits
        CHUNK = 10_000
        for i in range(0, len(rows_to_write), CHUNK):
            chunk = rows_to_write[i : i + CHUNK]
            ws.append_rows(chunk, value_input_option="RAW")

        rows_written = len(df)
        return {
            "rows_written": rows_written,
            "spreadsheet_url": spreadsheet_url,
            "sheet_name": sheet_name,
        }
